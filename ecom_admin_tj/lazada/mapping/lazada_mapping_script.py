import pandas as pd
import warnings
import os

# Suppress openpyxl warning about default style
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def load_lazada_product_to_dataframe(path: str) -> pd.DataFrame:
    dtype = {'Product ID': str, 'ชื่อสินค้า': str}
    lazada_products_df = pd.read_excel(
        path, sheet_name="template", usecols=['Product ID', 'ชื่อสินค้า'], dtype=dtype, engine='openpyxl')
    lazada_products_df = lazada_products_df[3:]
    lazada_products_df.reset_index(drop=True, inplace=True)
    # change column names
    lazada_products_df.columns = ['item_id', 'item_name']
    return lazada_products_df

def create_item_mapping_excel():
    """
    Create an Excel file with stock items, platform items, and mapping template.
    
    The Excel file will contain:
    - Sheet 1: Stock Items (from stock_items.csv)
    - Sheet 2: Platform Items (from platform_items.csv)
    - Sheet 3: Item Mapping (template for mapping - 1 platform item can have multiple stock items)
    
    Note: One platform item can be mapped to multiple stock items with different multipliers.
    """
    
    # Initialize configuration
    stock_items_path = 'ecom_admin_tj/common/stock_items.csv'
    # json_path = 'ecom_admin_tj/tiktok/mapping/tiktok_product_local_products_list.json'
    platform_items_path = 'ecom_admin_tj/lazada/mapping/lazada_products.xlsx'
    output_file = 'ecom_admin_tj/lazada/mapping/lazada_item_mapping.xlsx'
    
    # Check if input files exist
    if not os.path.exists(stock_items_path):
        raise FileNotFoundError(f"Stock items file not found: {stock_items_path}")
    if not os.path.exists(platform_items_path):
        raise FileNotFoundError(f"Platform items file not found: {platform_items_path}")

    
    print(f"Reading stock items from: {stock_items_path}")
    stock_df = pd.read_csv(stock_items_path)
    print(f"  Found {len(stock_df)} stock items")
    
    print(f"\nReading platform items from: {platform_items_path}")
    platform_df = load_lazada_product_to_dataframe(platform_items_path)
    print(f"  Found {len(platform_df)} platform items")
    
    # Create mapping template with empty rows for flexible mapping
    # Users can add multiple rows for the same platform_item_id to map to different stock items
    # Start with more rows than platform items to allow multiple mappings per platform item
    num_rows = len(platform_df) * 3  # Allow up to 3 mappings per platform item on average
    
    mapping_template = pd.DataFrame({
        'platform_item_id': [''] * num_rows,
        'platform_item_name': [''] * num_rows,
        'stock_item_id': [''] * num_rows,
        'stock_item_name': [''] * num_rows,
        'multiplier': [1.0] * num_rows  # Default multiplier is 1.0
    })
    
    # Create Excel writer
    print(f"\nCreating Excel file: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write Stock Items sheet
        stock_df.to_excel(writer, sheet_name='Stock Items', index=False)
        
        # Write Platform Items sheet
        platform_df.to_excel(writer, sheet_name='Platform Items', index=False)
        
        # Write Item Mapping sheet (template)
        mapping_template.to_excel(writer, sheet_name='Item Mapping', index=False)
        
        # Get the workbook and sheets to format
        # workbook = writer.book
        
        # Format Stock Items sheet
        stock_sheet = writer.sheets['Stock Items']
        stock_sheet.column_dimensions['A'].width = 15  # item_id
        stock_sheet.column_dimensions['B'].width = 60  # item_name
        
        # Format Platform Items sheet
        platform_sheet = writer.sheets['Platform Items']
        platform_sheet.column_dimensions['A'].width = 15  # item_id
        platform_sheet.column_dimensions['B'].width = 80  # item_name
        platform_sheet.column_dimensions['C'].width = 20  # item_sku
        platform_sheet.column_dimensions['D'].width = 15  # item_status
        platform_sheet.column_dimensions['E'].width = 12  # has_model
        platform_sheet.column_dimensions['F'].width = 12  # model_count
        platform_sheet.column_dimensions['G'].width = 12  # image_count
        
        # Format Item Mapping sheet
        mapping_sheet = writer.sheets['Item Mapping']
        mapping_sheet.column_dimensions['A'].width = 20  # platform_item_id
        mapping_sheet.column_dimensions['B'].width = 80  # platform_item_name
        mapping_sheet.column_dimensions['C'].width = 20  # stock_item_id
        mapping_sheet.column_dimensions['D'].width = 60  # stock_item_name
        mapping_sheet.column_dimensions['E'].width = 15  # multiplier
        
        # Add data validation and formulas
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Add instruction row at the top with styling
        mapping_sheet.insert_rows(1)
        mapping_sheet['A1'] = '📝 Instructions: One platform item can have multiple rows for different stock items. Fill platform_item_id, select stock_item_id from dropdown, and set multiplier.'
        mapping_sheet['A1'].font = Font(bold=True, color='FFFFFF')
        mapping_sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        mapping_sheet['A1'].alignment = Alignment(wrap_text=True)
        mapping_sheet.merge_cells('A1:E1')
        mapping_sheet.row_dimensions[1].height = 30
        
        # Create dropdown for platform_item_id column (Column A)
        platform_item_ids = platform_df['item_id'].astype(str).tolist()
        dv_platform = DataValidation(
            type="list",
            formula1=f'"' + ','.join(platform_item_ids[:100]) + '"',  # Limit to first 100 items
            allow_blank=True
        )
        dv_platform.error = 'Please select a valid platform item ID'
        dv_platform.errorTitle = 'Invalid Item ID'
        mapping_sheet.add_data_validation(dv_platform)
        dv_platform.add(f'A3:A{num_rows + 2}')  # Apply to column A (platform_item_id), skip instruction row
        
        # Add VLOOKUP formula for platform_item_name (Column B)
        # Formula will lookup platform_item_id in Platform Items sheet and return item_name
        for row in range(3, num_rows + 3):  # Start from row 3 (after instruction and header)
            mapping_sheet[f'B{row}'] = f'=IFERROR(VLOOKUP(A{row},\'Platform Items\'!A:B,2,FALSE),"")'
        
        # Create dropdown for stock_item_id column (Column C)
        stock_item_ids = stock_df['item_id'].astype(str).tolist()
        dv_stock = DataValidation(
            type="list",
            formula1=f'"' + ','.join(stock_item_ids) + '"',
            allow_blank=True
        )
        dv_stock.error = 'Please select a valid stock item ID'
        dv_stock.errorTitle = 'Invalid Item ID'
        mapping_sheet.add_data_validation(dv_stock)
        dv_stock.add(f'C3:C{num_rows + 2}')  # Apply to column C (stock_item_id), skip instruction row
        
        # Add VLOOKUP formula for stock_item_name (Column D)
        # Formula will lookup stock_item_id in Stock Items sheet and return item_name
        for row in range(3, num_rows + 3):  # Start from row 3
            mapping_sheet[f'D{row}'] = f'=IFERROR(VLOOKUP(C{row},\'Stock Items\'!A:B,2,FALSE),"")'
    
    print(f"\n✅ Excel file created successfully!")
    print(f"   Location: {output_file}")
    print(f"\n📋 File structure:")
    print(f"   - Sheet 1: Stock Items ({len(stock_df)} items)")
    print(f"   - Sheet 2: Platform Items ({len(platform_df)} items)")
    print(f"   - Sheet 3: Item Mapping ({num_rows} empty rows for flexible mapping)")
    print(f"\n💡 Next steps:")
    print(f"   1. Open the Excel file")
    print(f"   2. Go to 'Item Mapping' sheet")
    print(f"   3. For each platform item:")
    print(f"      - Select platform_item_id from dropdown (Column A)")
    print(f"      - platform_item_name will auto-fill (Column B)")
    print(f"      - Select stock_item_id from dropdown (Column C)")
    print(f"      - stock_item_name will auto-fill (Column D)")
    print(f"      - Set multiplier (Column E, default: 1.0)")
    print(f"   4. You can add MULTIPLE rows for the same platform_item_id")
    print(f"      to map it to different stock items")
    print(f"   5. Example: Gift Set A → Candy × 2, Toothpaste × 1, Mouthwash × 1")
    print(f"   6. Save the file for further processing")
    print(f"\n⚠️  Note: One platform item can have multiple s   tock item mappings")
    
    return output_file

if __name__ == "__main__":
    create_item_mapping_excel()
