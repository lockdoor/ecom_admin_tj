from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet

class ExcelFormatMixin:
    def _formating_header(
        self, sheet: Worksheet, row_height: int | None = None, font_color: str='FFFFFF', font_size: int=16, 
        start_color: str='4472C4', end_color: str='4472C4', fill_type: str='solid',
        horizontal: str='center', vertical: str='top', wrap_text: bool=True) -> None:
        """Apply header formatting to given sheet
        Args:
            sheet: Worksheet to format
            font_color: Font color for header
            font_size: Font size for header
            start_color: Fill start color for header
            end_color: Fill end color for header
            fill_type: Fill type for header
            horizontal: Horizontal alignment
            vertical: Vertical alignment
            wrap_text: Whether to wrap text
        """
        if row_height is not None:
            sheet.row_dimensions[1].height = row_height
        for cell in sheet[1]:
            cell.font = Font(bold=True, color=font_color, size=font_size)
            cell.fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
            cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
            
    def _formatting_body(
        self, sheet: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int,
        row_height: int=24, font_color: str='000000', font_size: int=14) -> None:
        """Apply body formatting to given sheet
        Args:
            sheet: Worksheet to format
            start_row: Starting row number for body formatting
            end_row: Ending row number for body formatting
            start_col: Starting column number for body formatting
            end_col: Ending column number for body formatting
            row_height: Height of each row
            font_color: Font color for body
            font_size: Font size for body
        """
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            sheet.row_dimensions[row[0].row].height = row_height
            for cell in row:
                cell.font = Font(color=font_color, size=font_size)
                
    def _formatting_footer(
        self, sheet: Worksheet, footer_row: int, row_height: int | None = None, 
        font_color: str='FFFFFF', font_size: int=16,
        start_color: str='4472C4', end_color: str='4472C4', fill_type: str='solid',
        vertical: str='top', wrap_text: bool=True) -> None:
        """Apply footer formatting to given sheet
        Args:
            sheet: Worksheet to format
            footer_row: Row number of the footer
            row_height: Height of the footer row
            font_color: Font color for footer
            font_size: Font size for footer
            start_color: Fill start color for footer
            end_color: Fill end color for footer
            fill_type: Fill type for footer
            vertical: Vertical alignment
            wrap_text: Whether to wrap text
        """
        if row_height is not None:
            sheet.row_dimensions[footer_row].height = row_height
        for cell in sheet[footer_row]:
            cell.font = Font(bold=True, color=font_color, size=font_size)
            cell.fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
            cell.alignment = Alignment(vertical=vertical, wrap_text=wrap_text)
