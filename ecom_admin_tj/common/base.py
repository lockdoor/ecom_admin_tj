from abc import ABC, abstractmethod
import pandas as pd
from pathlib import Path
import argparse
from datetime import date, datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet


class Base(ABC):
    SCRIPT_DIR: Path | None = None
    MAPPING_FILE: Path | None = None
    ORIGINAL_SHEET_NAME: str | None = None
    
    def __init__(self, input_file: str, output_file: str = None, shipping_date: datetime = None):
        """
        Initialize with input file path
        
        Args:
            input_file: Path to the input Excel file
            output_file: Optional custom output file path
            date: Optional date for filtering/processing
        """
        self.input_file: str = input_file
        self.output_file: str = output_file or self._generate_output_filename()
        self.shipping_date: datetime | None = shipping_date
        self.mapping_df: pd.DataFrame | None = None
        self.original_df: pd.DataFrame | None = None
        self.main_df: pd.DataFrame | None = None
        self.finance_df: pd.DataFrame | None = None
        self.merged_df: pd.DataFrame | None = None
        self.invoice_df: pd.DataFrame | None = None
        self.canceled_orders_df: pd.DataFrame | None = None
        self.order_sn_unique: int = 0
        self.merge_left : str | None = None
        self.merge_right : str | None = None
    
    def _generate_output_filename(self) -> str:
        """Generate output filename from input filename"""
        if self.input_file.endswith('_output.xlsx'):
            return self.input_file
        return self.input_file.replace('.xlsx', '_output.xlsx')
    
    @abstractmethod
    def load_mapping(self) -> pd.DataFrame:
        """Load item mapping from Excel file"""
        pass
    
    @abstractmethod
    def load_main_df(self) -> pd.DataFrame:
        """Load main data from input file"""
        pass

    @abstractmethod
    def calculate_finance_df(self) -> pd.DataFrame:
        """Calculate finance dataframe from main_df dataframe"""
        pass

    def load_canceled_orders(self) -> pd.DataFrame:
        """Load canceled orders from input file if exists"""
        if self.canceled_orders_df is None:
            try :
                self.canceled_orders_df = pd.read_excel(self.input_file, sheet_name='canceled_orders', dtype={'canceled_orders_sn': str})
            # ValueError occurs when sheet does not exist
            except (ValueError):
                print('No canceled orders sheet found. Continuing without excluding any orders.')
                self.canceled_orders_df = pd.DataFrame(columns=['canceled_orders_sn'], dtype=str)
        return self.canceled_orders_df
    
    def merge_mapping(self) -> pd.DataFrame:
        """Merge main dataframe with mapping"""
        
        if self.merge_left is None or self.merge_right is None:
            raise ValueError("merge_left and merge_right attributes must be set before merging.")
        
        self.merged_df = pd.merge(
            self.main_df, 
            self.mapping_df, 
            left_on=self.merge_left, 
            right_on=self.merge_right, 
            how='left')
        return self.merged_df
    
    @abstractmethod
    def calculate_invoice(self) -> pd.DataFrame:
        """Calculate invoice from merged dataframe"""
        pass
    
    @abstractmethod
    def export_excel(self) -> None:
        """Export invoice to Excel file"""
        pass
    
    def _cancel_orders_to_excel(self, writer: pd.ExcelWriter) -> None:
        """Export canceled orders to Excel with formatting"""
        self.canceled_orders_df.to_excel(writer, sheet_name='canceled_orders', index=False)
        canceled_sheet = writer.sheets['canceled_orders']
        canceled_sheet.column_dimensions['A'].width = 25  # canceled_orders_sn
        self._formating_header(
            canceled_sheet, row_height=30, font_color='FFFFFF', font_size=16,
            start_color='FF0000', end_color='FF0000', fill_type='solid',
            horizontal='center', vertical='center', wrap_text=True)
        self._formatting_body(
            canceled_sheet, start_row=2, end_row=200, start_col=1, end_col=1,
            row_height=24, font_color='FF0000', font_size=14)
                
    def _formating_header(
        self, sheet: Worksheet, row_height: int | None = None, font_color: str='FFFFFF', font_size: int=16, 
        start_color: str='4472C4', end_color: str='4472C4', fill_type: str='solid',
        horizontal: str='center', vertical: str='top', wrap_text: bool=True) -> None:
        """Apply header formatting to given sheet
        Args:
            sheet: Worksheet to format
            font_color: Font color for header
            font_size: Font size for header
            start_color: Fill start color for header
            end_color: Fill end color for header
            fill_type: Fill type for header
            horizontal: Horizontal alignment
            vertical: Vertical alignment
            wrap_text: Whether to wrap text
        """
        if row_height is not None:
            sheet.row_dimensions[1].height = row_height
        for cell in sheet[1]:
            cell.font = Font(bold=True, color=font_color, size=font_size)
            cell.fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
            cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
            
    def _formatting_body(
        self, sheet: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int,
        row_height: int=24, font_color: str='000000', font_size: int=14) -> None:
        """Apply body formatting to given sheet
        Args:
            sheet: Worksheet to format
            start_row: Starting row number for body formatting
            end_row: Ending row number for body formatting
            start_col: Starting column number for body formatting
            end_col: Ending column number for body formatting
            row_height: Height of each row
            font_color: Font color for body
            font_size: Font size for body
        """
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            sheet.row_dimensions[row[0].row].height = row_height
            for cell in row:
                cell.font = Font(color=font_color, size=font_size)
                
    def _formatting_footer(
        self, sheet: Worksheet, footer_row: int, row_height: int | None = None, 
        font_color: str='FFFFFF', font_size: int=16,
        start_color: str='4472C4', end_color: str='4472C4', fill_type: str='solid',
        vertical: str='top', wrap_text: bool=True) -> None:
        """Apply footer formatting to given sheet
        Args:
            sheet: Worksheet to format
            footer_row: Row number of the footer
            row_height: Height of the footer row
            font_color: Font color for footer
            font_size: Font size for footer
            start_color: Fill start color for footer
            end_color: Fill end color for footer
            fill_type: Fill type for footer
            vertical: Vertical alignment
            wrap_text: Whether to wrap text
        """
        if row_height is not None:
            sheet.row_dimensions[footer_row].height = row_height
        for cell in sheet[footer_row]:
            cell.font = Font(bold=True, color=font_color, size=font_size)
            cell.fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
            cell.alignment = Alignment(vertical=vertical, wrap_text=wrap_text)
    
    def process(self) -> None:
        """Main execution flow - template method pattern"""
        print(f"Starting {self.__class__.__name__}...")
        print(f'Reading input file: {self.input_file}')
        print(f'Processing date: {self.shipping_date.strftime("%Y-%m-%d") if self.shipping_date else "Not specified"}')
        
        # Load data
        self.mapping_df = self.load_mapping()
        self.main_df = self.load_main_df()
        
        # Process
        self.merged_df = self.merge_mapping()
        self.invoice_df = self.calculate_invoice()
        self.finance_df = self.calculate_finance_df()
        
        print(f'Unique order numbers processed: {self.order_sn_unique}')
        
        # Export
        print(f'Exporting to Excel file: {self.output_file}')
        self.export_excel()
        
        print("Process completed successfully!")
    
    @classmethod
    def create_argument_parser(cls) -> argparse.ArgumentParser:
        """
        Create argument parser with common arguments
        Subclasses can override to add custom arguments
        
        Returns:
            Configured ArgumentParser instance
        """
        platform_name = cls.__module__.split('.')[-2]
        
        parser = argparse.ArgumentParser(
            description=f'Process {platform_name.capitalize()} orders and generate invoice',
            formatter_class=argparse.RawDescriptionHelpFormatter
        )
        
        parser.add_argument(
            'input_file',
            type=str,
            help='Path to input Excel file'
        )
        
        parser.add_argument(
            '-o', '--output',
            type=str,
            dest='output_file',
            help='Path to output Excel file (default: input_file_output.xlsx)'
        )
        
        parser.add_argument(
            '-d', '--shipping_date',
            type=str,
            dest='shipping_date',
            help='Processing date in YYYY-MM-DD format (default: None)'
        )
        
        return parser
    
    @classmethod
    def from_args(cls, args: list[str] = None):
        """
        Factory method to create instance from command line arguments
        
        Args:
            args: Command line arguments (defaults to sys.argv[1:])
        
        Returns:
            Instance of the class
        """
        parser = cls.create_argument_parser()
        parsed_args = parser.parse_args(args)
        
        # Parse date if provided
        shipping_date = None
        if parsed_args.shipping_date:
            try:
                shipping_date = datetime.strptime(parsed_args.shipping_date, '%Y-%m-%d')
            except ValueError:
                parser.error(f'Invalid date format: {parsed_args.shipping_date}. Use YYYY-MM-DD')
        
        return cls(
            input_file=parsed_args.input_file,
            output_file=parsed_args.output_file,
            shipping_date=shipping_date
        )
