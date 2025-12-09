from ..common.base import Base
import pandas as pd
import numpy as np
from pathlib import Path

class Tiktok(Base):
    
    SCRIPT_DIR = Path(__file__).parent
    MAPPING_FILE = SCRIPT_DIR / 'tiktok_item_mapping.xlsx'
    ORIGINAL_SHEET_NAME = 'OrderSKUList'
    
    def __init__(self, input_file: str, output_file: str = None, shipping_date = None, mapping_file: str = None):
        """Initialize Tiktok processor with specific settings
                
        Args:
            input_file: Path to input Excel file
            output_file: Optional custom output file path
            shipping_date: Optional date for filtering/processing
        """
        # Pass None for shipping_date since Lazada doesn't use it
        if shipping_date is not None:
            print('Warning: shipping_date parameter is not used in Lazada processing.')
        super().__init__(input_file, output_file, shipping_date = None, mapping_file=mapping_file)
        
        # Set Tiktok-specific attributes
        self.SCRIPT_DIR = Path(__file__).parent
        if self.MAPPING_FILE is None:
            self.MAPPING_FILE = self.SCRIPT_DIR / "tiktok_item_mapping.xlsx"
        self.ORIGINAL_SHEET_NAME = "OrderSKUList"
        self.merge_left = 'SKU ID'
        self.merge_right = 'platform_item_id'
        
    def load_mapping(self) -> pd.DataFrame:
        """Load item mapping specific to Tiktok"""
        mapping_file_path = self.MAPPING_FILE
        mapping_type_dict = {
            'platform_item_id': str,
            'platform_item_name': str,
            'stock_item_id': str,
            'stock_item_name': str,
            'multiplier': np.int64,
        }
        self.mapping_df = pd.read_excel(mapping_file_path, sheet_name='Item Mapping', skiprows=1, dtype=mapping_type_dict)
        self.mapping_df.dropna(subset=['platform_item_id'], inplace=True)
        return self.mapping_df
    
    def load_main_df(self) -> pd.DataFrame:
        """Load main data from Tiktok input file"""
        type_dict = {
            'Order ID': str,
            'SKU ID': str,
            'Quantity': np.int64,
            'SKU Unit Original Price': np.float64,
            'SKU Subtotal Before Discount': np.float64,
            'SKU Seller Discount': np.float64,
            'SKU Subtotal After Discount': np.float64,
            }
        
        self.original_df = pd.read_excel(
            self.input_file, 
            sheet_name=self.ORIGINAL_SHEET_NAME, 
            dtype=type_dict, header=0, 
            skiprows=[1])
        
        if "Cancelation/Return Type" not in self.original_df.columns:
            # ถ้าอ่านด้วย pandas ไม่เจอคอลัมน์ "Cancelation/Return Type"
            # อ่านข้อมูลดิบจาก openpyxl แล้วแปลงเป็น DataFrame เอง
            from openpyxl import load_workbook

            wb = load_workbook(self.input_file, read_only=True, data_only=True)
            ws = wb.active

            # อ่าน header จากแถวที่ 1
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                headers.append(header)

            # print(f"Headers ({len(headers)} columns):")
            # print(headers[:10])

            # อ่านข้อมูลเริ่มจากแถวที่ 3 (ข้าม header และ description)
            data = []
            for row in range(3, ws.max_row + 1):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                data.append(row_data)

            # สร้าง DataFrame
            self.original_df = pd.DataFrame(data, columns=headers)
            self.original_df = self.original_df.astype(type_dict)
        
        df = self.original_df.copy()
        
        # clean dataframe
        df = df[df["Cancelation/Return Type"].isna()]
        df.reset_index(inplace=True)
        
        columns= ['Order ID', 'SKU ID', 'Product Name', 'Quantity', 'SKU Unit Original Price', 'SKU Subtotal Before Discount', 'SKU Seller Discount', 'SKU Subtotal After Discount']
        df = df[columns]

        # read canceled sheets
        self.load_canceled_orders()
        canceled_order_sns = self.canceled_orders_df['canceled_orders_sn'].dropna().unique()
        df = df[~df['Order ID'].isin(canceled_order_sns)]
        
        # count unique order numbers
        self.order_sn_unique = df['Order ID'].nunique()

        self.main_df = df
        return self.main_df
    
    def merge_mapping(self) -> pd.DataFrame:
        """Merge main dataframe with mapping"""
        super().merge_mapping()
        self.merged_df['จำนวนรวม'] = self.merged_df['Quantity'] * self.merged_df['multiplier']
        return self.merged_df
    
    def calculate_invoice(self):
        
        if self.merged_df is None:
            raise ValueError("Merged dataframe is not available. Please run merge_mapping() first.")
        
        self.invoice_df = self.merged_df.groupby('stock_item_id').agg({
        'stock_item_name': 'first',
        'จำนวนรวม': 'sum',
        'SKU Subtotal Before Discount': 'sum',
        'SKU Seller Discount': 'sum'
        }).reset_index()
        self.invoice_df.loc['TOTAL'] = [
            'TOTAL',
            '', 
            '', 
            self.invoice_df['SKU Subtotal Before Discount'].sum(), 
            self.invoice_df['SKU Seller Discount'].sum()]
        return self.invoice_df
    
    def calculate_finance_df(self) -> pd.DataFrame:
        """Calculate finance dataframe from main_df dataframe"""
        if self.merged_df is None:
            raise ValueError("Merged dataframe is not available. Please run merge_mapping() first.")
        
        self.finance_df = self.merged_df.groupby('Order ID', sort=False).agg({
            'SKU Subtotal Before Discount': 'sum',
            'SKU Seller Discount': 'sum',
            'SKU Subtotal After Discount': 'sum',   
        }).reset_index()
        
        # Add footer row with totals
        total_row = {
            'Order ID': 'TOTAL',
            'SKU Subtotal Before Discount': self.finance_df['SKU Subtotal Before Discount'].sum(),
            'SKU Seller Discount': self.finance_df['SKU Seller Discount'].sum(),
            'SKU Subtotal After Discount': self.finance_df['SKU Subtotal After Discount'].sum(),
        }
        self.finance_df.loc[len(self.finance_df)] = total_row
        
        return self.finance_df
    
    def export_excel(self):
        """Export Tiktok invoice to Excel file"""
        
        from openpyxl.worksheet.worksheet import Worksheet
        
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Sheet 1: Original orders 
            self.original_df.to_excel(writer, sheet_name=self.ORIGINAL_SHEET_NAME, index=False)
            original_sheet: Worksheet = writer.sheets[self.ORIGINAL_SHEET_NAME]
            self._formating_header(original_sheet)
            
            # Sheet 2: invoice
            self.invoice_df.to_excel(writer, sheet_name=f'invoice_{self.order_sn_unique}_orders', index=False)
            invoice_sheet: Worksheet = writer.sheets[f'invoice_{self.order_sn_unique}_orders']
            invoice_sheet.column_dimensions['A'].width = 18  # stock_item_id
            invoice_sheet.column_dimensions['B'].width = 48  # stock_item_name
            invoice_sheet.column_dimensions['C'].width = 14  # จำนวนรวม
            invoice_sheet.column_dimensions['D'].width = 14  # SKU Subtotal Before Discount
            invoice_sheet.column_dimensions['E'].width = 14  # SKU Seller Discount
            self._formating_header(sheet=invoice_sheet)
            self._formatting_body(sheet=invoice_sheet, start_row=2, end_row=len(self.invoice_df), start_col=1, end_col=5)
            self._formatting_footer(sheet=invoice_sheet, footer_row=len(self.invoice_df)+1)
            
            # Canceled orders (ensure string format)
            self.canceled_orders_df.to_excel(writer, sheet_name='canceled_orders', index=False)
            self._cancel_orders_to_excel(writer)
            
            # Finance summary
            self.finance_df.to_excel(writer, sheet_name='Finance Summary', index=False)
            finance_sheet: Worksheet = writer.sheets['Finance Summary']
            finance_sheet.column_dimensions['A'].width = 25  # Order ID
            finance_sheet.column_dimensions['B'].width = 18  # SKU Subtotal Before Discount
            finance_sheet.column_dimensions['C'].width = 18  # SKU Seller Discount
            finance_sheet.column_dimensions['D'].width = 18  # SKU Subtotal After Discount
            self._formating_header(finance_sheet)
            self._formatting_body(sheet=finance_sheet, start_row=2, end_row=len(self.finance_df), start_col=1, end_col=4)
            self._formatting_footer(sheet=finance_sheet, footer_row=len(self.finance_df)+1)
